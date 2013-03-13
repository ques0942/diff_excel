# -*- coding: utf-8


import openpyxl as xl


def diff_wb(wb1, wb2):
    """
    workbookの比較を行うためのメソッド．
    workbookがもつ各シートをそれぞれ比較して，各シートごとの比較結果を表示する．
    """
    sheet_names1 = wb1.get_sheet_names()
    sheet_names2 = wb2.get_sheet_names()
    diff_sheets = []
    result_diff = {}
    for sheet_name in sheet_names1[:]:
        if sheet_name in sheet_names2:
            ws1 = wb1.get_sheet_by_name(sheet_name)
            ws2 = wb2.get_sheet_by_name(sheet_name)
            result_diff[sheet_name] = diff_ws(ws1, ws2)
            sheet_names1.remove(sheet_name)
            sheet_names2.remove(sheet_name)
        else:
            diff_sheets.append(sheet_name)
    if sheet_names2:
        diff_sheets.extend(sheet_names2)

    return diff_sheets, result_diff


def diff_ws(ws1, ws2):
    """
    worksheetの比較を行うためのメソッド．
    """
    import difflib
    dim1 = ws1.calculate_dimension()
    dim2 = ws2.calculate_dimension()
    dim1 = ws1.range(dim1)
    dim2 = ws2.range(dim2)
    csv1 = []
    csv2 = []
    for row in dim1:
        tmp = []
        for col in row:
            tmp.append((col.value))
        csv1.append(tuple(tmp))
    for row in dim2:
        tmp = []
        for col in row:
            tmp.append((col.value))
        csv2.append(tuple(tmp))

    d = difflib.Differ()
    return list(d.compare(csv1, csv2))

def convert_csv(book_name, num=0):
    """
    xlsxのシートのうち1つをcsvに変換する．
    デフォルトでは，先頭のシートを変換する．
    """
    wb = xl.load_workbook(book_name)
    sheet_names = wb.get_sheet_names()
    target_name = sheet_names[num]
    ws = wb.get_sheet_by_name(target_name)
    dim = ws.calculate_dimension()
    dim = ws.range(dim)
    csv = []
    for row in dim:
        tmp = []
        for col in row:
            tmp.append(str(col.value))
        csv.append(tmp)
    return csv

if __name__ == '__main__':
    import sys
    from pprint import pprint
    if len(sys.argv) < 2:
        print 'no input'
        sys.exit(-1)
    elif len(sys.argv) == 2:
        for row in convert_csv(sys.argv[1]):
            print ','.join(row)
    else:
        book_name1 = sys.argv[1]
        book_name2 = sys.argv[2]
        wb1 = xl.load_workbook(book_name1)
        wb2 = xl.load_workbook(book_name2)
        diff_sheets, result_diff = diff_wb(wb1, wb2)
        pprint(result_diff)
